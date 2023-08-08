import re
import string
import os
import glob

import PyPDF2
import fitz

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

days_of_the_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

excel_data = {
    'Date of Service': [],
    'Location': [],
    'Patient Name': [],
    'Medical Number': [],
    'Procedure': [],
    'CPT Codes': [],
    'Comments': [],
}

ub_data = {
    'Line 1': [],
    'Line 2': [],
    'Line 3': []
}


def extract_text(pdf_path):
    doc = fitz.open(pdf_path)
    page = doc[0]
    text = page.get_text("text")

    return text


def combine_pdf_pages(input_path):
    src = fitz.open(input_path)
    doc = fitz.open()  # empty output PDF

    width = src[0].rect.width  # width of the pages
    total_height = sum([page.rect.height for page in src])  # total height of all pages

    # Create a new page with the total width and height
    page = doc.new_page(-1, width=width, height=total_height)

    y_offset = 0  # offset where to insert the next page

    # Loop over each page in the original PDF
    for spage in src:
        # Calculate the rectangle where to insert the page
        r = fitz.Rect(0, y_offset, width, y_offset + spage.rect.height)
        # Insert the page
        page.show_pdf_page(r, src, spage.number)
        # Update the offset
        y_offset += spage.rect.height

    # Save the new PDF
    doc.save(input_path, garbage=3, deflate=True)


def initialize_data():
    pdf_files = glob.glob(os.path.join('files', '*.pdf'))

    for pdf in pdf_files:
        combine_pdf_pages(pdf)

    return pdf_files


def format_lines(lines):
    lines = [''.join(ch for ch in line if ch in string.printable) for line in lines]
    lines = [line for line in lines if line.strip()]
    return lines


def check_line(line):
    pattern = r"^(?!Location:)[\W]*.*//.*//.*//?.*"
    match = re.search(pattern, line.strip())
    return match


def check_urgent_board(line):
    if "URGENT BOARD" in line:
        return True


def format_text(file_name):
    text = extract_text(f"{file_name}.pdf")
    lines = text.strip().split("\n")
    lines = format_lines(lines)

    events = []
    urgent_boards = []

    for index in range(len(lines)):
        line = lines[index]

        if check_urgent_board(line):
            urgent_boards.append([lines[index], lines[index + 1], lines[index + 2]])
            continue

        if check_line(line):
            event = []
            for x in range(index + 3, len(lines)):
                pattern = r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun) (\d{1,2}/\d{1,2}/\d{4}) (\d{1,2}:\d{2} (AM|PM)) (?:-|to) ((Mon|Tue|Wed|Thu|Fri|Sat|Sun)? \d{1,2}/\d{1,2}/\d{4} )?(\d{1,2}:\d{2} (AM|PM))$'
                match = re.match(pattern, lines[x])
                if match:
                    break
                if "Location: " in lines[x - 2]:
                    event = ['', '', '', '']
                    event[0] = line
                    event[1] = lines[x - 3].strip()
                    event[2] = lines[x - 2].strip()
                    try:
                        if not any(day in lines[x] for day in days_of_the_week):
                            pattern = r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun) (\d{1,2}/\d{1,2}/\d{4}) (\d{1,2}:\d{2} (AM|PM)) (?:-|to) ((Mon|Tue|Wed|Thu|Fri|Sat|Sun)? \d{1,2}/\d{1,2}/\d{4} )?(\d{1,2}:\d{2} (AM|PM))$'
                            run = True
                            count = 0
                            while run:
                                match = re.match(pattern, lines[x + 1 + count].strip())
                                if not match:
                                    if not any(day in lines[x + count] for day in days_of_the_week):
                                        event[3] += lines[x + count].strip() + ' || '
                                    else:
                                        run = False
                                else:
                                    run = False
                                count += 1

                        break
                    except Exception as e:
                        try:
                            for a in range(x, len(lines)):
                                event[3] += lines[a].strip()
                            print("file finished")
                        except Exception as e:
                            print("file finished")

            if event:
                events.append(event)

    return events, urgent_boards


def format_events(events, df):
    for event in events:
        # print (event)
        comments = ''.join(char for char in event[3] if char in string.printable)

        # First section of array
        current = event[0].split("//")
        location = current[0].strip()
        medical_number = current[1].strip()
        patient_name = current[2].strip()

        # Second section of array
        current = event[1]
        pattern = r'\d+/\d+/\d+'
        match = re.search(pattern, current)
        date = (match.group()).strip()

        # Third section of array
        current = event[2].split("//")
        procedure = (current[0][len("Location: "):]).strip()

        new_row = pd.DataFrame({
            'Date of Service': [date],
            'Location': [location],
            'Patient Name': [patient_name],
            'Medical Number': [medical_number],
            'Procedure': [procedure],
            'CPT Codes': [''],
            'Comments': [comments]
        })
        df = pd.concat([df, new_row], ignore_index=True)
    return df


def create_excel(df, df2, file_name):
    workbook = Workbook()

    # Remove the default sheet created by openpyxl
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Create sheets and populate them with data
    sheet1 = workbook.create_sheet(title="Main Calendar")
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet1.append(row)

    sheet2 = workbook.create_sheet(title="Urgent Board")
    for row in dataframe_to_rows(df2, index=False, header=True):
        sheet2.append(row)

    # Save the workbook
    workbook.save(f"output/{file_name}.xlsx")


def format_ub(urgent_boards, df2):
    # Create a new row using a DataFrame
    for event in urgent_boards:
        new_row = pd.DataFrame({
            'Line 1': [event[0]],
            'Line 2': [event[1]],
            'Line 3': [event[2]]
        })

        # Concatenate the new row to the existing DataFrame
        df2 = pd.concat([df2, new_row], ignore_index=True)

    return df2


def resize_columns(file_path):
    # Load the workbook
    book = load_workbook(f"output/{file_path}.xlsx")

    # Iterate through all the sheets in the workbook
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]

        # Iterate through the columns
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]

            # Check the length of each cell in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set the width of the column based on the max_length
            adjusted_width = max_length
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the workbook with the updated column sizes
    book.save(f"output/{file_path}.xlsx")


def delete_files_in_folder(folder_path):
    # Get a list of all files in the folder
    file_list = os.listdir(folder_path)

    # Iterate through the list and delete each file
    for file_name in file_list:
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)
            print(f"Deleted file: {file_name}")
        else:
            print(f"Skipping non-file item: {file_name}")


def clean_files():
    delete_files_in_folder("files")
    delete_files_in_folder("output/files")


def main():
    pdf_files = initialize_data()

    for pdf in pdf_files:
        df = pd.DataFrame(excel_data)
        df2 = pd.DataFrame(ub_data)

        file_name = os.path.splitext(pdf)[0]
        print(file_name)

        sorted_data = format_text(file_name)
        events = sorted_data[0]
        urgent_boards = sorted_data[1]

        df = format_events(events, df)
        df2 = format_ub(urgent_boards, df2)

        create_excel(df, df2, file_name)
        resize_columns(file_name)
